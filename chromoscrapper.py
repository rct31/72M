import time
import requests
import json
import os
from openpyxl import Workbook

url72m="    " #insérer ici l'ip de l'analyseur

urlchroma= url72m+"/cgi-bin/cgi_get_chroma_res"

page = requests.get(urlchroma)
Jsonchroma = json.loads(page.text) # contient les données du chromato

#création de la feuille excel
def extraction(sheet):
    page = requests.get(urlchroma)
    Jsonchroma = json.loads(page.text)    

    sheet.cell(row=2, column=1, value=url72m)


    start = Jsonchroma["gene"]["start"]
    lig1 = [start,"Nom","N°","rt","h_val","begin","end","m_height","area","conc","unit","d_tr","base_a","base_b","base_t"]
    
    for col, valeur in enumerate(lig1, start=1):
     sheet.cell(row=1, column=col, value=valeur)

    for i in range(1,len(Jsonchroma["comp"])+1): # DECOMPOSTION DU JSON + inscription dans le wb
     name= Jsonchroma["comp"]["C{}".format(i)]["name"]
     if name == 'CMP0057' :
      name = 'Benzène'
     elif name == 'CMP0058':
      name = 'Ethylbenzène'
     elif name == 'CMP0059':
      name = 'Toluène'
     elif name == 'CMP0060':
      name = 'm+p-xylene'
     elif name == 'CMP0061':
      name ="o-xylene"

     sheet.cell(row=i+1, column=2, value=name)

     n_pic = Jsonchroma["comp"]["C{}".format(i)]["n_pic"]
     sheet.cell(row=i+1, column=3, value=n_pic)

     rt = Jsonchroma["comp"]["C{}".format(i)]["rt"]
     sheet.cell(row=i+1, column=4, value=rt)

     h_val = Jsonchroma["comp"]["C{}".format(i)]["h_val"]
     sheet.cell(row=i+1, column=5, value=h_val)

     begin = Jsonchroma["comp"]["C{}".format(i)]["begin"]
     sheet.cell(row=i+1, column=6, value=begin)

     end = Jsonchroma["comp"]["C{}".format(i)]["end"]
     sheet.cell(row=i+1, column=7, value=end)

     m_height = Jsonchroma["comp"]["C{}".format(i)]["m_height"]
     sheet.cell(row=i+1, column=8, value=m_height)

     area = Jsonchroma["comp"]["C{}".format(i)]["area"]
     sheet.cell(row=i+1, column=9, value=area)

     conc = Jsonchroma["comp"]["C{}".format(i)]["conc"]
     sheet.cell(row=i+1, column=10, value=conc)

     unit = Jsonchroma["comp"]["C{}".format(i)]["unit"]
     if unit == "UNI0031":
       unit = "ppb"
     sheet.cell(row=i+1, column=11, value=unit)  
     
     d_tr = Jsonchroma["comp"]["C{}".format(i)]["d_tr"]
     sheet.cell(row=i+1, column=12, value=d_tr)

     base_a = Jsonchroma["comp"]["C{}".format(i)]["base_a"]
     sheet.cell(row=i+1, column=13, value=base_a)

     base_b = Jsonchroma["comp"]["C{}".format(i)]["base_b"]
     sheet.cell(row=i+1, column=14, value=base_b)

     base_t = Jsonchroma["comp"]["C{}".format(i)]["base_t"]
     sheet.cell(row=i+1, column=15, value=base_t)
  
     
    return start 

 



n = int(input("Entrez le nombre de chromatographes que vous souhaitez importer :"))
wb = Workbook()

current_start = None
 
if n == 1 :
 sheet = wb.active
 start = extraction(sheet) # heure de départ
 print(f"Chromatographe acquis. Heure de départ : {start}")
 sheet.title = "Chromato "

for i in range(1, n + 1): 
  if i== 1 :
    sheet = wb.active
    start = extraction(sheet) # heure de départ
    sheet.title = "Chromato 1 "
    

  else : 
    sheet = wb.create_sheet(title="Chromato " + str(i))
    start = extraction(sheet) # heure de départ
    
  
  if current_start is not None and start == current_start:
     while start == current_start : # attente du prochain QH 
      time.sleep(100)
      start = extraction(sheet)
      
  print(f"Chromatographe {i} acquis. Heure de départ : {start}")

  current_start = start

#sauvegarde du fichier excel   
script_path = os.path.abspath(__file__) #chemin absolu du script
script_directory = os.path.dirname(script_path) # nom du dossier du script
excel_file_path = os.path.join(script_directory, 'Chromatographe.xlsx')  # nom du ficher excel


wb.save(filename=excel_file_path)


